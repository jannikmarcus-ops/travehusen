from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Styles ---
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
subheader_font = Font(name='Arial', bold=True, size=10, color='2F5496')
subheader_fill = PatternFill('solid', fgColor='D6E4F0')
data_font = Font(name='Arial', size=10)
input_font = Font(name='Arial', size=10, color='0000FF')
input_fill = PatternFill('solid', fgColor='FFFFCC')
title_font = Font(name='Arial', bold=True, size=14, color='2F5496')
section_font = Font(name='Arial', bold=True, size=12, color='2F5496')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
euro_format = '#,##0.00 €'
qm_format = '#,##0.00'
qm_price_format = '#,##0.00 €'

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, max_col, is_input_cols=None):
    if is_input_cols is None:
        is_input_cols = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if col in is_input_cols:
            cell.font = input_font
            cell.fill = input_fill

def style_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border

# ============================================================
# SHEET 1: Übersicht alle Einheiten
# ============================================================
ws = wb.active
ws.title = 'Übersicht'

headers = [
    'Nr.', 'Bezeichnung', 'Gebäude / Baufeld', 'Typ', 'Zimmer',
    'Geschosse', 'Wohnfläche (m²)', 'Grundstück (m²)',
    'Terrasse/Balkon (m²)', 'Besonderheiten',
    'Verkaufspreis (€)', 'Preis/m² (€)', 'Bemerkungen'
]
col_widths = [6, 16, 20, 22, 10, 14, 18, 18, 20, 35, 18, 14, 25]

ws.cell(row=1, column=1, value='Travehusen - Preiskalkulation').font = title_font
ws.merge_cells('A1:M1')
ws.cell(row=2, column=1, value='Übersicht aller Einheiten').font = section_font
ws.merge_cells('A2:M2')

header_row = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=header_row, column=i, value=h)
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws, header_row, len(headers))

# Input columns: Verkaufspreis (11), Bemerkungen (13)
input_cols = [11, 13]
row = header_row + 1

# --- BF 12.1 Einzelhäuser ---
ws.cell(row=row, column=1, value='Baufeld 12.1 - Einzelhäuser')
ws.merge_cells(f'A{row}:M{row}')
style_subheader_row(ws, row, len(headers))
row += 1

bf12_1 = [
    ('1', 'Hs. 12.1-01', 'BF 12.1', 'Einzelhaus', '-', '-', None, 448, None, 'Keine Detailpläne vorliegend'),
    ('2', 'Hs. 12.1-02', 'BF 12.1', 'Einzelhaus', '-', '-', None, 502, None, 'inkl. Weg'),
    ('3', 'Hs. 12.1-03', 'BF 12.1', 'Einzelhaus', '-', '-', None, 557, None, 'inkl. Weg'),
    ('4', 'Hs. 12.1-04', 'BF 12.1', 'Einzelhaus', '-', '-', None, 448, None, ''),
    ('5', 'Hs. 12.1-05', 'BF 12.1', 'Einzelhaus', '-', '-', None, 448, None, ''),
    ('6', 'Hs. 12.1-06', 'BF 12.1', 'Einzelhaus', '-', '-', None, 448, None, ''),
    ('7', 'Hs. 12.1-07', 'BF 12.1', 'Einzelhaus', '-', '-', None, 448, None, ''),
]
for d in bf12_1:
    for i, v in enumerate(d, 1):
        ws.cell(row=row, column=i, value=v)
    # Preis/m² Formel
    ws.cell(row=row, column=12, value=f'=IF(AND(K{row}<>"",G{row}<>""),K{row}/G{row},"")')
    ws.cell(row=row, column=12).number_format = qm_price_format
    ws.cell(row=row, column=11).number_format = euro_format
    if d[6]: ws.cell(row=row, column=7).number_format = qm_format
    style_data_row(ws, row, len(headers), input_cols)
    row += 1

# --- BF 12.2 Reihenhäuser ---
ws.cell(row=row, column=1, value='Baufeld 12.2 - Reihenhäuser')
ws.merge_cells(f'A{row}:M{row}')
style_subheader_row(ws, row, len(headers))
row += 1

bf12_2 = [
    ('1', 'Hs. 12.2-01', 'BF 12.2', 'ERH mit Einlieger-WE', '3+2', 'EG+OG+SG', 128.67, 241, 11.25, 'Hauptwhg. 83,94 m² + Einlieger 44,73 m², Terrasse, Dachterrasse'),
    ('2', 'Hs. 12.2-02', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('3', 'Hs. 12.2-03', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('4', 'Hs. 12.2-04', 'BF 12.2', 'ERH 5,00m', '4-5', 'EG+OG+SG', 134.40, 241, 11.47, 'Terrasse, Dachterrasse, Galerie, Garten'),
    ('5', 'Hs. 12.2-05', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('6', 'Hs. 12.2-06', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('7', 'Hs. 12.2-07', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('8', 'Hs. 12.2-08', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('9', 'Hs. 12.2-09', 'BF 12.2', 'ERH 5,00m', '4-5', 'EG+OG+SG', 134.40, 241, 11.47, 'Terrasse, Dachterrasse, Galerie, Garten'),
    ('10', 'Hs. 12.2-10', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('11', 'Hs. 12.2-11', 'BF 12.2', 'MRH 6,00m', '4', 'EG+OG', 99.71, 174, 6.00, 'Terrasse, Garten'),
    ('12', 'Hs. 12.2-12', 'BF 12.2', 'ERH 5,00m', '4-5', 'EG+OG+SG', 134.40, 241, 11.47, 'Terrasse, Dachterrasse, Galerie, Garten'),
]
for d in bf12_2:
    for i, v in enumerate(d, 1):
        ws.cell(row=row, column=i, value=v)
    ws.cell(row=row, column=12, value=f'=IF(AND(K{row}<>"",G{row}<>""),K{row}/G{row},"")')
    ws.cell(row=row, column=12).number_format = qm_price_format
    ws.cell(row=row, column=11).number_format = euro_format
    ws.cell(row=row, column=7).number_format = qm_format
    style_data_row(ws, row, len(headers), input_cols)
    row += 1

# --- BF 12.3 Doppelhaushälften ---
ws.cell(row=row, column=1, value='Baufeld 12.3 - Doppelhaushälften')
ws.merge_cells(f'A{row}:M{row}')
style_subheader_row(ws, row, len(headers))
row += 1

bf12_3 = [
    ('1', 'Hs. 12.3-01', 'BF 12.3', 'DHH Variante 1', '4-5', 'EG+OG+SB', 111.49, 275, 7.50, 'Terrasse, Spitzboden, Garten, Stellplatz'),
    ('2', 'Hs. 12.3-02', 'BF 12.3', 'DHH Variante 1', '4-5', 'EG+OG+SB', 111.49, 268, 7.50, 'Terrasse, Spitzboden, Garten, Stellplatz'),
    ('3', 'Hs. 12.3-03', 'BF 12.3', 'DHH Variante 2', '4-5', 'EG+OG+SB', 111.49, 268, 7.50, 'Terrasse, Spitzboden, Garten'),
    ('4', 'Hs. 12.3-04', 'BF 12.3', 'DHH Variante 2', '4-5', 'EG+OG+SB', 111.49, 268, 7.50, 'Terrasse, Spitzboden, Garten'),
    ('5', 'Hs. 12.3-05', 'BF 12.3', 'DHH Variante 2', '4-5', 'EG+OG+SB', 111.49, 268, 7.50, 'Terrasse, Spitzboden, Garten'),
    ('6', 'Hs. 12.3-06', 'BF 12.3', 'DHH Variante 2', '4-5', 'EG+OG+SB', 111.49, 268, 7.50, 'Terrasse, Spitzboden, Garten'),
    ('7', 'Hs. 12.3-07', 'BF 12.3', 'DHH Variante 2', '4-5', 'EG+OG+SB', 111.49, 268, 7.50, 'Terrasse, Spitzboden, Garten'),
    ('8', 'Hs. 12.3-08', 'BF 12.3', 'DHH Variante 2', '4-5', 'EG+OG+SB', 111.49, 268, 7.50, 'Terrasse, Spitzboden, Garten'),
]
for d in bf12_3:
    for i, v in enumerate(d, 1):
        ws.cell(row=row, column=i, value=v)
    ws.cell(row=row, column=12, value=f'=IF(AND(K{row}<>"",G{row}<>""),K{row}/G{row},"")')
    ws.cell(row=row, column=12).number_format = qm_price_format
    ws.cell(row=row, column=11).number_format = euro_format
    ws.cell(row=row, column=7).number_format = qm_format
    style_data_row(ws, row, len(headers), input_cols)
    row += 1

# --- BF 12.4 Reihenhäuser Tiny ---
ws.cell(row=row, column=1, value='Baufeld 12.4 - Reihenhäuser (Tiny-House)')
ws.merge_cells(f'A{row}:M{row}')
style_subheader_row(ws, row, len(headers))
row += 1

bf12_4 = [
    ('1', 'Hs. 12.4-01', 'BF 12.4', 'ERH 5,00m', '4-5', 'EG+OG+SG', 134.40, 246, 11.47, 'Terrasse, Dachterrasse, Garten'),
    ('2', 'Hs. 12.4-02', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('3', 'Hs. 12.4-03', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('4', 'Hs. 12.4-04', 'BF 12.4', 'ERH 5,00m (Treppe gedreht)', '5', 'EG+OG+SG', 133.36, 251, 11.47, 'Terrasse, Dachterrasse, Garten'),
    ('5', 'Hs. 12.4-05', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('6', 'Hs. 12.4-06', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('7', 'Hs. 12.4-07', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('8', 'Hs. 12.4-08', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('9', 'Hs. 12.4-09', 'BF 12.4', 'ERH 5,00m', '4-5', 'EG+OG+SG', 134.40, 243, 11.47, 'Terrasse, Dachterrasse, Garten'),
    ('10', 'Hs. 12.4-10', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('11', 'Hs. 12.4-11', 'BF 12.4', 'MRH Tiny-House 5,10m', '3', 'EG+OG', 90.17, 154, 6.00, 'Terrasse, Garten'),
    ('12', 'Hs. 12.4-12', 'BF 12.4', 'ERH 5,00m', '4-5', 'EG+OG+SG', 134.40, 251, 11.47, 'Terrasse, Dachterrasse, Garten'),
]
for d in bf12_4:
    for i, v in enumerate(d, 1):
        ws.cell(row=row, column=i, value=v)
    ws.cell(row=row, column=12, value=f'=IF(AND(K{row}<>"",G{row}<>""),K{row}/G{row},"")')
    ws.cell(row=row, column=12).number_format = qm_price_format
    ws.cell(row=row, column=11).number_format = euro_format
    ws.cell(row=row, column=7).number_format = qm_format
    style_data_row(ws, row, len(headers), input_cols)
    row += 1

# --- BF 12.5 Doppelhaushälften ---
ws.cell(row=row, column=1, value='Baufeld 12.5 - Doppelhaushälften')
ws.merge_cells(f'A{row}:M{row}')
style_subheader_row(ws, row, len(headers))
row += 1

bf12_5 = [
    ('1', 'Hs. 12.5-01', 'BF 12.5', 'DHH Variante 1', '4-5', 'EG+OG+SB', 120.38, 277, 7.50, 'Terrasse, Spitzboden, Garten, Stellplatz'),
    ('2', 'Hs. 12.5-02', 'BF 12.5', 'DHH Variante 1', '4-5', 'EG+OG+SB', 120.38, 277, 7.50, 'Terrasse, Spitzboden, Garten, Stellplatz'),
    ('3', 'Hs. 12.5-03', 'BF 12.5', 'DHH Variante 2', '4-5', 'EG+OG+SB', 119.84, 277, 7.50, 'Terrasse, Spitzboden, Garten, opt. Homeoffice'),
    ('4', 'Hs. 12.5-04', 'BF 12.5', 'DHH Variante 2', '4-5', 'EG+OG+SB', 119.84, 277, 7.50, 'Terrasse, Spitzboden, Garten, opt. Homeoffice'),
]
for d in bf12_5:
    for i, v in enumerate(d, 1):
        ws.cell(row=row, column=i, value=v)
    ws.cell(row=row, column=12, value=f'=IF(AND(K{row}<>"",G{row}<>""),K{row}/G{row},"")')
    ws.cell(row=row, column=12).number_format = qm_price_format
    ws.cell(row=row, column=11).number_format = euro_format
    ws.cell(row=row, column=7).number_format = qm_format
    style_data_row(ws, row, len(headers), input_cols)
    row += 1

# ============================================================
# SHEET 2: Haus 3 - Wohnungen
# ============================================================
ws2 = wb.create_sheet('Haus 3')

headers2 = [
    'Nr.', 'WE-Nr.', 'Geschoss', 'Typ', 'Zimmer',
    'Wohnfläche (m²)', 'Terrasse/Balkon brutto (m²)', 'Terrasse/Balkon anteilig (m²)',
    'Besonderheiten',
    'Verkaufspreis (€)', 'Preis/m² (€)', 'Bemerkungen'
]
col_widths2 = [6, 12, 12, 16, 10, 18, 24, 24, 40, 18, 14, 25]

ws2.cell(row=1, column=1, value='Travehusen - Haus 3 - Preiskalkulation').font = title_font
ws2.merge_cells('A1:L1')
ws2.cell(row=2, column=1, value='28 Wohneinheiten, barrierefrei, Tiefgarage mit 22 Stellplätzen').font = Font(name='Arial', size=10, italic=True, color='666666')
ws2.merge_cells('A2:L2')

header_row2 = 4
for i, h in enumerate(headers2, 1):
    ws2.cell(row=header_row2, column=i, value=h)
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws2, header_row2, len(headers2))

input_cols2 = [10, 12]
row2 = header_row2 + 1

haus3_data = {
    'Erdgeschoss': [
        (1, '301.0', 'EG', '1,5-Zimmer', '1,5', 46.70, 12.05, 6.02, 'Terrasse, barrierefrei'),
        (2, '302.0', 'EG', '3-Zimmer', '3', 76.11, 17.60, 8.80, 'Terrasse, barrierefrei'),
        (3, '303.0', 'EG', '3,5-Zimmer', '3,5', 91.90, 18.83, 9.41, 'Terrasse, barrierefrei'),
        (4, '304.0', 'EG', '2-Zimmer', '2', 75.17, 15.50, 7.75, 'Terrasse, barrierefrei'),
        (5, '316.0', 'EG', '2-Zimmer', '2', 75.18, 15.50, 7.75, 'Terrasse, barrierefrei'),
        (6, '317.0', 'EG', '3,5-Zimmer', '3,5', 91.91, 18.83, 9.42, 'Terrasse, barrierefrei'),
        (7, '318.0', 'EG', '3-Zimmer', '3', 76.11, 17.60, 8.80, 'Terrasse, barrierefrei'),
        (8, '319.0', 'EG', '1,5-Zimmer', '1,5', 46.70, 12.05, 6.03, 'Terrasse, barrierefrei'),
    ],
    '1. Obergeschoss': [
        (9, '305.1', '1. OG', '1,5-Zimmer', '1,5', 43.20, 7.70, 3.85, 'Balkon, barrierefrei'),
        (10, '306.1', '1. OG', '3-Zimmer', '3', 71.16, 7.70, 3.85, 'Balkon, barrierefrei'),
        (11, '307.1', '1. OG', '3,5-Zimmer', '3,5', 86.33, 7.70, 3.85, 'Balkon, barrierefrei'),
        (12, '308.1', '1. OG', '2-Zimmer', '2', 79.85, 7.70, 3.85, 'Balkon, barrierefrei'),
        (13, '320.1', '1. OG', '2-Zimmer', '2', 79.85, 7.70, 3.85, 'Balkon, barrierefrei'),
        (14, '321.1', '1. OG', '3,5-Zimmer', '3,5', 86.33, 7.70, 3.85, 'Balkon, barrierefrei'),
        (15, '322.1', '1. OG', '3-Zimmer', '3', 71.15, 7.70, 3.85, 'Balkon, barrierefrei'),
        (16, '323.1', '1. OG', '1,5-Zimmer', '1,5', 43.19, 7.70, 3.85, 'Balkon, barrierefrei'),
    ],
    '2. Obergeschoss': [
        (17, '309.2', '2. OG', '1,5-Zimmer', '1,5', 43.20, 7.70, 3.85, 'Balkon, barrierefrei'),
        (18, '310.2', '2. OG', '3-Zimmer', '3', 71.16, 7.70, 3.85, 'Balkon, barrierefrei'),
        (19, '311.2', '2. OG', '3,5-Zimmer', '3,5', 86.50, 7.70, 3.85, 'Balkon, barrierefrei'),
        (20, '312.2', '2. OG', '2-Zimmer', '2', 79.97, 7.70, 3.85, 'Balkon, barrierefrei'),
        (21, '324.2', '2. OG', '2-Zimmer', '2', 79.97, 7.70, 3.85, 'Balkon, barrierefrei'),
        (22, '325.2', '2. OG', '3,5-Zimmer', '3,5', 86.50, 7.70, 3.85, 'Balkon, barrierefrei'),
        (23, '326.2', '2. OG', '3-Zimmer', '3', 71.15, 7.70, 3.85, 'Balkon, barrierefrei'),
        (24, '327.2', '2. OG', '1,5-Zimmer', '1,5', 43.20, 7.70, 3.85, 'Balkon, barrierefrei'),
    ],
    '3. Obergeschoss (Staffel)': [
        (25, '313.3', '3. OG', '1,5-Zimmer', '1,5', 43.20, 7.70, 3.85, 'Balkon, barrierefrei'),
        (26, '314.3', '3. OG', '3-Zimmer', '3', 71.17, 7.70, 3.85, 'Balkon, barrierefrei'),
        (27, '315.3', '3. OG', '2-Zimmer', '2', 98.92, 32.69, 16.35, 'Dachterrasse + Balkon, Arbeits-/Gästezimmer, WC extra, Premium'),
        (28, '328.3', '3. OG', '2-Zimmer', '2', 98.79, 32.70, 16.35, 'Dachterrasse + Balkon, Arbeits-/Gästezimmer, WC extra, Premium'),
        (29, '329.3', '3. OG', '3-Zimmer', '3', 71.17, 7.70, 3.85, 'Balkon, barrierefrei'),
        (30, '330.3', '3. OG', '1,5-Zimmer', '1,5', 43.20, 7.70, 3.85, 'Balkon, barrierefrei'),
    ],
}

for section, units in haus3_data.items():
    ws2.cell(row=row2, column=1, value=section)
    ws2.merge_cells(f'A{row2}:L{row2}')
    style_subheader_row(ws2, row2, len(headers2))
    row2 += 1
    for d in units:
        for i, v in enumerate(d, 1):
            ws2.cell(row=row2, column=i, value=v)
        ws2.cell(row=row2, column=11, value=f'=IF(AND(J{row2}<>"",F{row2}<>""),J{row2}/F{row2},"")')
        ws2.cell(row=row2, column=11).number_format = qm_price_format
        ws2.cell(row=row2, column=10).number_format = euro_format
        ws2.cell(row=row2, column=6).number_format = qm_format
        ws2.cell(row=row2, column=7).number_format = qm_format
        ws2.cell(row=row2, column=8).number_format = qm_format
        style_data_row(ws2, row2, len(headers2), input_cols2)
        row2 += 1

# ============================================================
# SHEET 3: Zusammenfassung nach Typ
# ============================================================
ws3 = wb.create_sheet('Zusammenfassung')

ws3.cell(row=1, column=1, value='Travehusen - Zusammenfassung nach Gebäudetyp').font = title_font
ws3.merge_cells('A1:G1')

headers3 = ['Gebäude / Baufeld', 'Typ', 'Anzahl', 'Wohnfläche (m²)', 'Grundstück (m²)', 'Durchschnittspreis (€)', 'Durchschnitts-m²-Preis (€)']
col_widths3 = [22, 28, 10, 18, 18, 22, 24]

for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws3, 3, len(headers3))

summary_data = [
    ('BF 12.1', 'Einzelhaus', 7, '-', '448-557', '', ''),
    ('BF 12.2', 'MRH 6,00m (4-Zi.)', 8, '99,71', '174', '', ''),
    ('BF 12.2', 'ERH 5,00m (4-5-Zi.)', 3, '134,40', '241', '', ''),
    ('BF 12.2', 'ERH mit Einlieger-WE', 1, '128,67', '241', '', ''),
    ('BF 12.3', 'DHH Variante 1 (6,80x11,00m)', 2, '111,49', '268-275', '', ''),
    ('BF 12.3', 'DHH Variante 2 (6,80x11,00m)', 6, '111,49', '268', '', ''),
    ('BF 12.4', 'MRH Tiny-House 5,10m (3-Zi.)', 8, '90,17', '154', '', ''),
    ('BF 12.4', 'ERH 5,00m (4-5-Zi.)', 4, '133,36-134,40', '243-251', '', ''),
    ('BF 12.5', 'DHH Variante 1 (6,80x11,80m)', 2, '120,38', '277', '', ''),
    ('BF 12.5', 'DHH Variante 2 (6,80x11,80m)', 2, '119,84', '277', '', ''),
    ('Haus 3', '1,5-Zimmer-Wohnung', 8, '43,19-46,70', '-', '', ''),
    ('Haus 3', '2-Zimmer-Wohnung', 4, '75,17-79,97', '-', '', ''),
    ('Haus 3', '2-Zimmer-Wohnung (Premium, 3.OG)', 2, '98,79-98,92', '-', '', ''),
    ('Haus 3', '3-Zimmer-Wohnung', 8, '71,15-76,11', '-', '', ''),
    ('Haus 3', '3,5-Zimmer-Wohnung', 6, '86,33-91,91', '-', '', ''),
]

for idx, d in enumerate(summary_data):
    r = 4 + idx
    for i, v in enumerate(d, 1):
        ws3.cell(row=r, column=i, value=v)
    ws3.cell(row=r, column=6).number_format = euro_format
    ws3.cell(row=r, column=7).number_format = qm_price_format
    style_data_row(ws3, r, len(headers3), [6, 7])

# Gesamt-Zeile
total_row = 4 + len(summary_data) + 1
ws3.cell(row=total_row, column=1, value='GESAMT')
ws3.cell(row=total_row, column=3, value=f'=SUM(C4:C{total_row-2})')
for col in range(1, len(headers3) + 1):
    ws3.cell(row=total_row, column=col).font = Font(name='Arial', bold=True, size=10)
    ws3.cell(row=total_row, column=col).border = Border(top=Side(style='double'), bottom=Side(style='double'))

# Freeze panes
ws.freeze_panes = 'A5'
ws2.freeze_panes = 'A5'
ws3.freeze_panes = 'A4'

# AutoFilter
ws.auto_filter.ref = f'A4:M{row-1}'
ws2.auto_filter.ref = f'A4:L{row2-1}'

# Print setup
for sheet in [ws, ws2, ws3]:
    sheet.sheet_properties.pageSetUpPr = None
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1

output = '/Users/jannikmarcus/Projects/travehusen/Travehusen_Preiskalkulation.xlsx'
wb.save(output)
print(f'Gespeichert: {output}')
