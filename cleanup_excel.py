from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from copy import copy

wb = load_workbook('/Users/jannikmarcus/Projects/travehusen/Travehusen_Preiskalkulation.xlsx')

# === SHEET: Übersicht ===
ws = wb['Übersicht']

# Sweet Spot prices mapping: row -> (Verkaufspreis, Wohnfläche or None for Grundstück-based)
# BF 12.1 Einzelhäuser (rows 6-12): Keep original Gemini prices as Verkaufspreis, no Wohnfläche
bf121_prices = {
    6: (394000, None, 448),   # 12.1-01, 448m² Grundstück
    7: (427000, None, 502),   # 12.1-02, 502m²
    8: (473000, None, 557),   # 12.1-03, 557m²
    9: (394000, None, 448),   # 12.1-04, 448m²
    10: (394000, None, 448),  # 12.1-05, 448m²
    11: (394000, None, 448),  # 12.1-06, 448m²
    12: (394000, None, 448),  # 12.1-07, 448m²
}

# BF 12.2 Reihenhäuser (rows 14-25)
bf122_prices = {
    14: (559000, 128.67),  # ERH mit Einlieger
    15: (419000, 99.71),   # MRH 6,00m
    16: (419000, 99.71),
    17: (569000, 134.4),   # ERH 5,00m
    18: (419000, 99.71),
    19: (419000, 99.71),
    20: (419000, 99.71),
    21: (419000, 99.71),
    22: (569000, 134.4),   # ERH 5,00m
    23: (419000, 99.71),
    24: (419000, 99.71),
    25: (569000, 134.4),   # ERH 5,00m
}

# BF 12.3 DHH (rows 27-34)
bf123_prices = {
    27: (489000, 111.49),  # DHH V1
    28: (489000, 111.49),  # DHH V1
    29: (465000, 111.49),  # DHH V2
    30: (465000, 111.49),
    31: (465000, 111.49),
    32: (465000, 111.49),
    33: (465000, 111.49),
    34: (465000, 111.49),
}

# BF 12.4 Reihenhäuser Tiny (rows 36-47)
bf124_prices = {
    36: (569000, 134.4),    # ERH 5,00m
    37: (399000, 90.17),    # Tiny
    38: (399000, 90.17),
    39: (569000, 133.36),   # ERH Treppe gedreht
    40: (399000, 90.17),
    41: (399000, 90.17),
    42: (399000, 90.17),
    43: (399000, 90.17),
    44: (569000, 134.4),    # ERH 5,00m
    45: (399000, 90.17),
    46: (399000, 90.17),
    47: (569000, 134.4),    # ERH 5,00m
}

# BF 12.5 DHH (rows 49-52)
bf125_prices = {
    49: (519000, 120.38),  # DHH V1
    50: (519000, 120.38),  # DHH V1
    51: (509000, 119.84),  # DHH V2
    52: (509000, 119.84),  # DHH V2
}

# Fill BF 12.1 (Einzelhäuser - €/m² based on Grundstück)
for row, (vk, _, grundst) in bf121_prices.items():
    ws.cell(row=row, column=12).value = vk  # L = Verkaufspreis
    ws.cell(row=row, column=13).value = round(vk / grundst) if grundst else None  # M = €/m² Grundstück

# Fill BF 12.2-12.5
for prices_dict in [bf122_prices, bf123_prices, bf124_prices, bf125_prices]:
    for row, (vk, wfl) in prices_dict.items():
        ws.cell(row=row, column=12).value = vk
        ws.cell(row=row, column=13).value = round(vk / wfl) if wfl else None

# Delete Gemini columns (N=14, O=15) - delete O first then N
ws.delete_cols(15)  # O = €/m² Gemini KI
ws.delete_cols(14)  # N = Preisindikation Gemini KI

# === SHEET: Haus 3 ===
ws3 = wb['Haus 3']

# Relana's prices for Haus 3
relana_haus3 = {
    6: (355000, 46.7),    # WE 301
    7: (569000, 76.11),   # WE 302
    8: (689000, 91.9),    # WE 303
    9: (559000, 75.17),   # WE 304
    10: (539000, 75.18),  # WE 316
    11: (649000, 91.91),  # WE 317
    12: (539000, 76.11),  # WE 318
    13: (335000, 46.7),   # WE 319
    15: (309000, 43.2),   # WE 305.1
    16: (509000, 71.16),  # WE 306.1
    17: (615000, 86.33),  # WE 307.1
    18: (559000, 79.85),  # WE 308.1
    19: (539000, 79.85),  # WE 320.1
    20: (575000, 86.33),  # WE 321.1
    21: (475000, 71.15),  # WE 322.1
    22: (295000, 43.19),  # WE 323.1
    24: (315000, 43.2),   # WE 309.2
    25: (515000, 71.16),  # WE 310.2
    26: (625000, 86.5),   # WE 311.2
    27: (569000, 79.97),  # WE 312.2
    28: (549000, 79.97),  # WE 324.2
    29: (585000, 86.5),   # WE 325.2
    30: (485000, 71.15),  # WE 326.2
    31: (295000, 43.2),   # WE 327.2
    33: (365000, 43.2),   # WE 313.3
    34: (595000, 71.17),  # WE 314.3
    35: (835000, 98.92),  # WE 315.3
    36: (805000, 98.79),  # WE 328.3
    37: (559000, 71.17),  # WE 329.3
    38: (345000, 43.2),   # WE 330.3
}

for row, (vk, wfl) in relana_haus3.items():
    ws3.cell(row=row, column=11).value = vk  # K = Verkaufspreis
    ws3.cell(row=row, column=12).value = round(vk / wfl) if wfl else None  # L = Preis/m²

# Delete Gemini columns from Haus 3 (M=13, N=14)
ws3.delete_cols(14)  # N = €/m² Gemini KI
ws3.delete_cols(13)  # M = Preisindikation Gemini KI

# === SHEET: Zusammenfassung ===
wsz = wb['Zusammenfassung']

# Fill summary with calculated averages
summary_data = {
    4: (394000 + 427000 + 473000 + 394000*4, 7, None),  # BF 12.1 Einzelhaus - avg VK, count, avg €/m²
}

# BF 12.1: avg price
bf121_total = 394000 + 427000 + 473000 + 394000 + 394000 + 394000 + 394000
bf121_avg = round(bf121_total / 7)
bf121_avg_m2 = 879  # avg €/m² Grundstück

# BF 12.2 MRH: 8 × 419k
bf122_mrh_avg = 419000
bf122_mrh_m2 = round(419000 / 99.71)

# BF 12.2 ERH: 3 × 569k
bf122_erh_avg = 569000
bf122_erh_m2 = round(569000 / 134.4)

# BF 12.2 Einlieger: 1 × 559k
bf122_einl_avg = 559000
bf122_einl_m2 = round(559000 / 128.67)

# BF 12.3 V1: 2 × 489k
bf123_v1_avg = 489000
bf123_v1_m2 = round(489000 / 111.49)

# BF 12.3 V2: 6 × 465k
bf123_v2_avg = 465000
bf123_v2_m2 = round(465000 / 111.49)

# BF 12.4 Tiny: 8 × 399k
bf124_tiny_avg = 399000
bf124_tiny_m2 = round(399000 / 90.17)

# BF 12.4 ERH: 4 units (3×569k ERH + 1×569k Treppe = all 569k)
bf124_erh_avg = 569000
bf124_erh_m2 = round(569000 / 134.1)  # avg of 134.4, 133.36, 134.4, 134.4

# BF 12.5 V1: 2 × 519k
bf125_v1_avg = 519000
bf125_v1_m2 = round(519000 / 120.38)

# BF 12.5 V2: 2 × 509k
bf125_v2_avg = 509000
bf125_v2_m2 = round(509000 / 119.84)

# Haus 3 categories
h3_15z = [355000, 335000, 309000, 295000, 315000, 295000, 365000, 345000]
h3_2z = [559000, 539000, 559000, 539000, 569000, 549000]  # regular 2-Zi (not premium)
h3_2z_prem = [835000, 805000]
h3_3z = [569000, 539000, 509000, 475000, 515000, 485000, 595000, 559000]
h3_35z = [689000, 649000, 615000, 575000, 625000, 585000]

# Haus 3 Wohnflächen
h3_15z_wfl = [46.7, 46.7, 43.2, 43.19, 43.2, 43.2, 43.2, 43.2]
h3_2z_wfl = [75.17, 75.18, 79.85, 79.85, 79.97, 79.97]
h3_2z_prem_wfl = [98.92, 98.79]
h3_3z_wfl = [76.11, 76.11, 71.16, 71.15, 71.16, 71.15, 71.17, 71.17]
h3_35z_wfl = [91.9, 91.91, 86.33, 86.33, 86.5, 86.5]

# Summary row data: (row, avg_price, avg_m2_price)
summary_rows = [
    (4, bf121_avg, bf121_avg_m2),
    (5, bf122_mrh_avg, bf122_mrh_m2),
    (6, bf122_erh_avg, bf122_erh_m2),
    (7, bf122_einl_avg, bf122_einl_m2),
    (8, bf123_v1_avg, bf123_v1_m2),
    (9, bf123_v2_avg, bf123_v2_m2),
    (10, bf124_tiny_avg, bf124_tiny_m2),
    (11, bf124_erh_avg, bf124_erh_m2),
    (12, bf125_v1_avg, bf125_v1_m2),
    (13, bf125_v2_avg, bf125_v2_m2),
    (14, round(sum(h3_15z)/len(h3_15z)), round(sum(p/w for p,w in zip(h3_15z, h3_15z_wfl))/len(h3_15z))),
    (15, round(sum(h3_2z)/len(h3_2z)), round(sum(p/w for p,w in zip(h3_2z, h3_2z_wfl))/len(h3_2z))),
    (16, round(sum(h3_2z_prem)/len(h3_2z_prem)), round(sum(p/w for p,w in zip(h3_2z_prem, h3_2z_prem_wfl))/len(h3_2z_prem))),
    (17, round(sum(h3_3z)/len(h3_3z)), round(sum(p/w for p,w in zip(h3_3z, h3_3z_wfl))/len(h3_3z))),
    (18, round(sum(h3_35z)/len(h3_35z)), round(sum(p/w for p,w in zip(h3_35z, h3_35z_wfl))/len(h3_35z))),
]

for row, avg_price, avg_m2 in summary_rows:
    wsz.cell(row=row, column=7).value = avg_price   # G = Durchschnittspreis
    wsz.cell(row=row, column=8).value = avg_m2       # H = Durchschnitts-m²-Preis

# Delete Gemini columns from Zusammenfassung (I=9, J=10)
wsz.delete_cols(10)  # J
wsz.delete_cols(9)   # I

# === FORMAT: Number formatting for currency ===
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# Übersicht: L (now col 12) = Verkaufspreis, M (now col 13) = €/m²
for row in range(6, ws.max_row + 1):
    cell_vk = ws.cell(row=row, column=12)
    cell_m2 = ws.cell(row=row, column=13)
    if cell_vk.value and isinstance(cell_vk.value, (int, float)):
        cell_vk.number_format = '#,##0 €'
    if cell_m2.value and isinstance(cell_m2.value, (int, float)):
        cell_m2.number_format = '#,##0'

# Haus 3: K (col 11) = Verkaufspreis, L (col 12) = €/m²
for row in range(6, ws3.max_row + 1):
    cell_vk = ws3.cell(row=row, column=11)
    cell_m2 = ws3.cell(row=row, column=12)
    if cell_vk.value and isinstance(cell_vk.value, (int, float)):
        cell_vk.number_format = '#,##0 €'
    if cell_m2.value and isinstance(cell_m2.value, (int, float)):
        cell_m2.number_format = '#,##0'

# Zusammenfassung: G (col 7) = Durchschnittspreis, H (col 8) = €/m²
for row in range(4, wsz.max_row + 1):
    cell_vk = wsz.cell(row=row, column=7)
    cell_m2 = wsz.cell(row=row, column=8)
    if cell_vk.value and isinstance(cell_vk.value, (int, float)):
        cell_vk.number_format = '#,##0 €'
    if cell_m2.value and isinstance(cell_m2.value, (int, float)):
        cell_m2.number_format = '#,##0'

# Save
output_path = '/Users/jannikmarcus/Projects/travehusen/Travehusen_Preiskalkulation_BPD.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
print('Done!')
